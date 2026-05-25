import logging
from celery import shared_task
from django.utils import timezone

logger = logging.getLogger(__name__)


@shared_task(bind=True, max_retries=3)
def generate_report(self, report_id):
    try:
        from apps.reports.models import Report
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        import io
        from django.core.files.base import ContentFile

        report = Report.objects.get(pk=report_id)
        report.status = 'processing'
        report.save(update_fields=['status'])

        wb = openpyxl.Workbook()
        ws = wb.active
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')

        if report.type == 'sales':
            _generate_sales_report(ws, report.parameters, header_font, header_fill)
        elif report.type == 'vendors':
            _generate_vendors_report(ws, report.parameters, header_font, header_fill)
        elif report.type == 'products':
            _generate_products_report(ws, report.parameters, header_font, header_fill)
        elif report.type == 'payments':
            _generate_payments_report(ws, report.parameters, header_font, header_fill)
        elif report.type == 'users':
            _generate_users_report(ws, report.parameters, header_font, header_fill)

        for col in ws.columns:
            max_length = max((len(str(cell.value or '')) for cell in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_length + 4, 50)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        filename = f'{report.type}_{report.pk}.xlsx'
        report.file.save(filename, ContentFile(buffer.read()), save=False)
        report.status = 'ready'
        report.completed_at = timezone.now()
        report.save(update_fields=['file', 'status', 'completed_at'])
        return {'status': 'success', 'report_id': report.pk}
    except Exception as exc:
        logger.error(f"Report generation error for {report_id}: {exc}")
        try:
            from apps.reports.models import Report
            Report.objects.filter(pk=report_id).update(status='failed')
        except Exception:
            pass
        raise self.retry(exc=exc, countdown=60)


def _apply_header(ws, headers, header_font, header_fill):
    ws.append(headers)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')


def _generate_sales_report(ws, params, header_font, header_fill):
    from apps.orders.models import Order
    ws.title = 'Ventes'
    _apply_header(ws, ['N° Commande', 'Client', 'Statut', 'Paiement', 'Montant', 'Date'], header_font, header_fill)
    qs = Order.objects.all().select_related('user').order_by('-created_at')
    date_from = params.get('date_from')
    date_to = params.get('date_to')
    if date_from:
        qs = qs.filter(created_at__date__gte=date_from)
    if date_to:
        qs = qs.filter(created_at__date__lte=date_to)
    for order in qs:
        ws.append([order.order_number, order.user.email, order.status, order.payment_status, float(order.total_amount), order.created_at.strftime('%Y-%m-%d %H:%M')])


def _generate_vendors_report(ws, params, header_font, header_fill):
    from apps.vendors.models import Store
    ws.title = 'Vendeurs'
    _apply_header(ws, ['Boutique', 'Propriétaire', 'Ville', 'Statut', 'Ventes', 'Revenus', 'Note', 'Date création'], header_font, header_fill)
    for store in Store.objects.all().select_related('user').order_by('-total_revenue'):
        ws.append([store.name, store.user.email, store.city, store.status, store.total_sales, float(store.total_revenue), float(store.rating), store.created_at.strftime('%Y-%m-%d')])


def _generate_products_report(ws, params, header_font, header_fill):
    from apps.products.models import Product
    ws.title = 'Produits'
    _apply_header(ws, ['Nom', 'Boutique', 'Catégorie', 'Prix', 'Stock', 'Note', 'Avis', 'Actif'], header_font, header_fill)
    for p in Product.objects.all().select_related('store', 'category').order_by('-created_at'):
        ws.append([p.name, p.store.name, p.category.name if p.category else '', float(p.price), p.stock, float(p.rating), p.reviews_count, p.is_active])


def _generate_payments_report(ws, params, header_font, header_fill):
    from apps.payments.models import Payment
    ws.title = 'Paiements'
    _apply_header(ws, ['Référence', 'Commande', 'Montant', 'Devise', 'Méthode', 'Passerelle', 'Statut', 'Date'], header_font, header_fill)
    for p in Payment.objects.all().select_related('order').order_by('-created_at'):
        ws.append([p.reference, p.order.order_number, float(p.amount), p.currency, p.method, p.gateway, p.status, p.created_at.strftime('%Y-%m-%d %H:%M')])


def _generate_users_report(ws, params, header_font, header_fill):
    from django.contrib.auth import get_user_model
    User = get_user_model()
    ws.title = 'Utilisateurs'
    _apply_header(ws, ['Email', 'Nom', 'Rôle', 'Téléphone', 'Email vérifié', 'Téléphone vérifié', 'Date inscription'], header_font, header_fill)
    for u in User.objects.all().order_by('-date_joined'):
        ws.append([u.email, u.get_full_name(), u.role, str(u.phone or ''), u.is_email_verified, u.is_phone_verified, u.date_joined.strftime('%Y-%m-%d')])
